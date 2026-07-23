# Copyright © 2026 OldNestling
# License: GPLv3 (GNU General Public License Version 3)

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook
import json, re
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.platypus.flowables import HRFlowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from Core.BoQ import BoQ_manager, Section, Work, Resource, Link
from Core.Project import Project
from Core.Utilities import text_after, clearing_string

from Templates.About import PROGRAM_NAME, PROGRAM_VERSION

class Convertor:

	# Имя фиктивного файла для работ без ссылок
	DUMMY_FILE_NAME = "_no_reference.pdf"
	DUMMY_PAGE_NUMBER = "1"
	DUMMY_DESCRIPTION = "Нет ссылки"

	def __init__(self, project):
		self.project: Project = project
		self.xml_dir = self.project.base_dir / 'XML'
		self.UNITS_LIB: dict | None = None	# Словарь с еденицами измерения
	
	def create_xml_3p01(self, manager: BoQ_manager, file_extension = '.gge') -> bool:
		project = self.project
		if not project.library:
			project.set_library()
		manager.set_links_file_id()


		# 1. Создаём корневой элемент с пространством имён
		root = ET.Element('Construction', {
			'xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance',
			'xsi:noNamespaceSchemaLocation': 'QuantityTakeoff-3_01.xsd'
		})

		# 2. Метаданные
		meta = ET.SubElement(root, 'Meta')
		soft = ET.SubElement(meta, 'Soft')
		ET.SubElement(soft, 'Name').text = PROGRAM_NAME
		ET.SubElement(soft, 'Version').text = PROGRAM_VERSION
		file_meta = ET.SubElement(meta, 'File')
		ET.SubElement(file_meta, 'Type').text = 'Ведомость объемов работ'
		ET.SubElement(file_meta, 'Version').text = '3.01'

		# 3. Строительная площадка и объект
		ET.SubElement(root, 'ConstructionSite').text = project.construction_site
		ET.SubElement(root, 'ObjectName').text = manager.object_name
		ET.SubElement(root, 'Num').text = manager.num
		ET.SubElement(root, 'Reason').text = manager.reason.replace('\n', '<br/>')

		# 4. Даты
		date = manager.date if manager.date else project.now.strftime("%d.%m.%Y")
		day, mounth, year = date.split('.')
		date = ET.SubElement(root, 'Date')
		ET.SubElement(date, 'Year').text = year
		ET.SubElement(date, 'Month').text = mounth
		ET.SubElement(date, 'Day').text = day

		# ISO-дата экспорта
		export_dt = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
		ET.SubElement(root, 'ExportDateTime').text = export_dt

		# 5. Подписи
		signatures = ET.SubElement(root, 'Signatures')
		composer = ET.SubElement(signatures, 'Composer')
		ET.SubElement(composer, 'Name').text = manager.signatures.get('Composer','')
		ET.SubElement(composer, 'Position').text = manager.signatures.get('Composer_Position','')

		if isinstance(manager.verifier, dict):
			verifier_data = manager.verifier
		else:
			verifier_data = self.project.verifier

		verifier = ET.SubElement(signatures, 'Verifier')
		ET.SubElement(verifier, 'Name').text = verifier_data.get('Name')
		ET.SubElement(verifier, 'Position').text = verifier_data.get('Position')

		# 6. Ссылки на файлы
		files = ET.SubElement(root, 'Files')
		all_files = manager.get_links_files()
		for file_data in all_files.values():
			file = ET.SubElement(files, 'File')
			ET.SubElement(file, 'ID').text = str(file_data.get('FileID'))
			full_link: Path = file_data.get('FullLink')
			ET.SubElement(file, 'FileName').text = full_link.name if full_link else file_data.get('FileName')

		# 7. Секции и работы
		sections_tag = ET.SubElement(root, 'Sections')
		manager.calculate_pos_nums()
		is_dummy_links = False
		for i, section in enumerate(manager.sections, 1):
			section: Section
			section_tag = ET.SubElement(sections_tag, 'Section')
			ET.SubElement(section_tag, 'Num').text = str(i)
			ET.SubElement(section_tag, 'Name').text = section.raw_name

			works_tag = ET.SubElement(section_tag, 'Works')
			for work in section.works:
				work: Work

				work_tag = ET.SubElement(works_tag, 'Work')

				ET.SubElement(work_tag, 'Num').text = str(work.num)
				ET.SubElement(work_tag, 'Type').text = work.type
				ET.SubElement(work_tag, 'Name').text = work.name
				ET.SubElement(work_tag, 'Unit').text = work.unit
				ET.SubElement(work_tag, 'Quantity').text = str(work.quantity)
				ET.SubElement(work_tag, 'QuantityFormula').text = str(work.quantity_formula)

				# Ссылка для работы
				links_tag = ET.SubElement(work_tag, 'Links')

				if work.links:
					for link in work.links:
						link: Link
						link_tag = ET.SubElement(links_tag, 'Link')
						ET.SubElement(link_tag, 'FileID').text = str(link.file_id)
						pages = link.pages
						if isinstance(pages, list):
							pages = ' '.join([str(page) for page in pages])
						else:
							pages = str(pages)
						ET.SubElement(link_tag, 'PageNumber').text = pages
						ET.SubElement(link_tag, 'PageDescription').text = str(link)
				else:
					if is_dummy_links is False:
						""" Добавляем заглушку в список файлов для позиций без ссылок """
						file = ET.SubElement(files, 'File')
						ET.SubElement(file, 'ID').text = str(len(all_files)+1)
						ET.SubElement(file, 'FileName').text = self.DUMMY_FILE_NAME
						all_files[0] = {'FileID': len(all_files)+1, 'FileName': self.DUMMY_FILE_NAME}
						is_dummy_links = True
					link_tag = ET.SubElement(links_tag, 'Link')
					ET.SubElement(link_tag, 'FileID').text = str(all_files[0]['FileID'])
					ET.SubElement(link_tag, 'PageNumber').text = str(self.DUMMY_PAGE_NUMBER)
					ET.SubElement(link_tag, 'PageDescription').text = self.DUMMY_DESCRIPTION

				ET.SubElement(work_tag, 'Comment').text = work.comment

				if work.resources:
					resources_tag = ET.SubElement(work_tag, 'Resources')
					for resource in work.resources:
						resource: Resource

						resource_tag = ET.SubElement(resources_tag, 'Resource')
						if manager.position_mode:
							num = text_after(str(resource.num), '.')
						else:
							num = str(resource.num)
						ET.SubElement(resource_tag, 'Num').text = num
						ET.SubElement(resource_tag, 'Type').text = resource.type
						ET.SubElement(resource_tag, 'Name').text = resource.name
						ET.SubElement(resource_tag, 'Unit').text = resource.unit
						ET.SubElement(resource_tag, 'Quantity').text = str(resource.quantity)
						ET.SubElement(resource_tag, 'QuantityFormula').text = str(resource.quantity_formula)

						# Ссылка для работы
						links_tag = ET.SubElement(resource_tag, 'Links')

						if resource.links:
							for link in resource.links:
								link: Link
								link_tag = ET.SubElement(links_tag, 'Link')
								ET.SubElement(link_tag, 'FileID').text = str(link.file_id)
								pages = link.pages
								if isinstance(pages, list):
									pages = ' '.join([str(page) for page in pages])
								else:
									pages = str(pages)
								ET.SubElement(link_tag, 'PageNumber').text = pages
								ET.SubElement(link_tag, 'PageDescription').text = str(link)
						else:
							if is_dummy_links is False:
								""" Добавляем заглушку в список файлов для позиций без ссылок """
								file = ET.SubElement(files, 'File')
								ET.SubElement(file, 'ID').text = str(len(all_files)+1)
								ET.SubElement(file, 'FileName').text = self.DUMMY_FILE_NAME
								all_files[0] = {'FileID': len(all_files)+1, 'FileName': self.DUMMY_FILE_NAME}
								is_dummy_links = True
							link_tag = ET.SubElement(links_tag, 'Link')
							ET.SubElement(link_tag, 'FileID').text = str(all_files[0]['FileID'])
							ET.SubElement(link_tag, 'PageNumber').text = str(self.DUMMY_PAGE_NUMBER)
							ET.SubElement(link_tag, 'PageDescription').text = self.DUMMY_DESCRIPTION
						
						ET.SubElement(resource_tag, 'Comment').text = resource.comment

		# 8. Запись в файл с отступами (pretty print)
		tree = ET.ElementTree(root)
		# Для красивого вывода можно использовать ET.indent (Python 3.9+)
		try:
			ET.indent(tree, space='	')
		except AttributeError:
			pass  # или использовать внешнюю библиотеку lxml
		folder = manager.file.parent.parent / 'XML'
		folder.mkdir(parents=True, exist_ok=True)
		output = folder / manager.file.stem 
		tree.write(output.with_suffix(file_extension), encoding='utf-8', xml_declaration=True)
		return True
	
	def import_from_xml_3p01(self, xml_path: Path):
		"""
		Создаёт JSON по схеме проекта из XML-файла формата QuantityTakeoff-3_01.
		Импортируются: разделы, работы, ресурсы, их имена, типы, единицы измерения,
		формулы объёмов и комментарии. Номера позиций, ссылки, Reason и Quantity игнорируются.
		"""
		def __finder_unit_key(unit):
			if self.UNITS_LIB is None:
				self.UNITS_LIB = self.project.units
			for key, data in self.UNITS_LIB.items():
				if data.get('label') == unit:
					return key
			compare = {
				'м3': 'cubic_meter',
				'м³ проф': 'cubic_meter_profile',
				'м3 проф': 'cubic_meter_profile',
				'м3 проф.': 'cubic_meter_profile',
				'м³ матер': 'cubic_meter_material',
				'м3 матер': 'cubic_meter_material',
				'м3 матер.': 'cubic_meter_material',
				'м³ констр': 'cubic_meter_constr',
				'м3 констр': 'cubic_meter_constr',
				'м3 констр.': 'cubic_meter_constr',
				'м³ грунт': 'cubic_meter_soil',
				'м3 грунт': 'cubic_meter_soil',
				'м3 грунт.': 'cubic_meter_soil',
				'шт': 'count',
				'уп': 'package',
				'п.м.': 'p_m'
			}
			for label, key in compare.items():
				if label == unit:
					return key
			return unit

		tree = ET.parse(xml_path)
		root = tree.getroot()

		meta = root.find('Meta')
		file = meta.find('File')
		file_type = file.find('Type')
		version = file.find('Version')

		if file_type.text is None or version is None:
			raise ShemaError(xml_path.name)
		if file_type.text != 'Ведомость объемов работ' or version.text != '3.01':
			raise ShemaError(xml_path.name, file_type.text, version.text)

		# Данные о разделе
		metadata = {}
		metadata['ObjectName'] = root.find('ObjectName').text
		metadata['Num'] = root.find('Num').text
		date = root.find('Date')
		metadata['Date'] = f'{date.findtext('Day')}.{date.findtext('Month')}.{date.findtext('Year')}'
		signatures = root.find('Signatures').find('Composer')
		metadata['Signatures'] = {
			'Composer': signatures.findtext('Name'),
			'Composer_Position': signatures.findtext('Position')
		}

		# Основные данные
		sections_elem = root.find('Sections')

		if sections_elem is None:
			return # TODO

		sections = []	# Накопитель всех разделов
		for section_elem in sections_elem.findall('Section'):
			section = {'name': section_elem.find('Name').text.replace('Раздел:','')}
			works = []

			works_elem = section_elem.find('Works')
			last_work = None	# ссылка на последнюю работу
			for work_elem in works_elem.findall('Work'):
				work = {}

				# Заполняем поля работы
				if (elem := work_elem.find('Name')) is not None and elem.text:
					work['raw_name'] = elem.text
				if (elem := work_elem.find('Unit')) is not None and elem.text:
					work['raw_unit'] = __finder_unit_key(elem.text)
				if (elem := work_elem.find('QuantityFormula')) is not None and elem.text:
					work['raw_quantity_formula'] = clearing_string(elem.text)
				if (elem := work_elem.find('Type')) is not None and elem.text:
					work['type'] = elem.text.lower()
				if (elem := work_elem.find('Comment')) is not None and elem.text:
					work['raw_comment'] = elem.text

				is_work = (work['type'] == 'работа')
				
				# Импорт ресурсов
				resources_elem = work_elem.find('Resources')

				# Если в схеме использовалась иерархия с Resources
				if resources_elem is not None:
					work['resources'] = []
					for res_elem in resources_elem.findall('Resource'):
						resource = {}
						if (elem := res_elem.find('Name')) is not None and elem.text:
							resource['raw_name'] = elem.text
						if (elem := res_elem.find('Unit')) is not None and elem.text:
							resource['raw_unit'] = __finder_unit_key(elem.text)
						if (elem := res_elem.find('QuantityFormula')) is not None and elem.text:
							resource['raw_quantity_formula'] = clearing_string(elem.text)
						if (elem := res_elem.find('Type')) is not None and elem.text:
							resource['type'] = elem.text.lower()
						if (elem := res_elem.find('Comment')) is not None and elem.text:
							resource['raw_comment'] = elem.text
						work['resources'].append(resource)
				# отметка последней позиции работы
			
				if is_work:
					last_work = work
					works.append(work)
				else:
					# Если это ресурс – прикрепляем к последней работе (если она есть)
					if last_work is not None:
						if 'resources' not in last_work:
							last_work['resources'] = []
						last_work['resources'].append(work)
					else:
						works.append(work)
			section['works'] = works
			sections.append(section)
		data = {'sections': sections, 'archive': list()}
		file_data = {'metadata': metadata, 'data': data}
		
		try:
			folder = self.project.project_BoQs_path
			folder.mkdir(parents=True, exist_ok=True)
			filename = xml_path.stem
			filepath = folder / f'{filename}.json'
			with open(filepath, "w", encoding="utf-8") as f:
				# Записываем словарь в файл с форматированием
				json.dump(file_data, f, indent=4, ensure_ascii=False)
		except Exception as e:
			print(f'[DEBAG: import_from_xml_3p01] Ошибка импорта: {e}')

	def import_from_excel(self, filepath: Path, sheet_names: list[str] = None) -> int:
		"""
		Импорт данных из Excel-файла ведомости (формат, аналогичный excel2xml.py).
		Для каждого выбранного листа создаётся JSON-файл в папке project_BoQs_path.
		
		:param filepath: путь к Excel-файлу
		:param sheet_names: список имён листов для обработки (если None, берутся все листы с A1="Документ")
		:return: количество успешно обработанных листов
		"""
		# --- Вспомогательные функции (адаптированы из excel2xml.py) ---
		def detect_section(row):
			"""Проверяет, является ли строка заголовком раздела."""
			for col_idx in range(3):  # A, B, C
				val = row[col_idx].value
				if val is None:
					continue
				s = str(val).strip()
				match = re.match(r'^\s*раздел\s*[:.\-]?\s*(.*)', s, re.IGNORECASE)
				if match:
					title = match.group(1).strip()
					return title if title else "Раздел без названия"
			return None

		def parse_work_number(value):
			"""Извлекает целочисленный номер работы из ячейки A."""
			if value is None:
				return None
			s = str(value).strip()
			if not s:
				return None
			s = re.sub(r'\s+', '', s)
			try:
				num = int(float(s))
				return num if num > 0 else None
			except ValueError:
				return None

		def normalize_page_numbers(value):
			"""Преобразует номера страниц в строку с пробелами, например '1 2 3'."""
			if value is None:
				return None
			s = str(value).strip()
			numbers = re.findall(r'\d+', s)
			if numbers:
				return ' '.join(str(int(num)) for num in numbers if int(num) > 0)
			return None

		def find_unit_key(unit_text):
			"""Преобразует текстовую единицу измерения в ключ из словаря units."""
			if not unit_text:
				return ""
			if self.UNITS_LIB is None:
				self.UNITS_LIB = self.project.units
			for key, data in self.UNITS_LIB.items():
				if data.get('label') == unit_text:
					return key
			# ручные соответствия
			compare = {
				'м3': 'cubic_meter',
				'м³ проф': 'cubic_meter_profile',
				'м3 проф': 'cubic_meter_profile',
				'м3 проф.': 'cubic_meter_profile',
				'м³ матер': 'cubic_meter_material',
				'м3 матер': 'cubic_meter_material',
				'м3 матер.': 'cubic_meter_material',
				'м³ констр': 'cubic_meter_constr',
				'м3 констр': 'cubic_meter_constr',
				'м3 констр.': 'cubic_meter_constr',
				'м³ грунт': 'cubic_meter_soil',
				'м3 грунт': 'cubic_meter_soil',
				'м3 грунт.': 'cubic_meter_soil',
				'шт': 'count',
				'уп': 'package',
				'п.м.': 'p_m'
			}
			for label, key in compare.items():
				if label == unit_text:
					return key
			return unit_text  # fallback

		def clean_string(val):
			"""Возвращает очищенную строку или пустую строку."""
			return str(val).strip() if val is not None else ""

		# --- Основная логика ---
		if not filepath.exists():
			raise FileNotFoundError(f"Файл не найден: {filepath}")

		wb = load_workbook(filepath, data_only=True)
		sheets_to_process = []
		if sheet_names:
			for name in sheet_names:
				if name in wb.sheetnames:
					sheets_to_process.append(name)
				else:
					print(f"Лист '{name}' не найден, пропуск")
		else:
			# Все листы, у которых A1 == "Документ"
			for name in wb.sheetnames:
				ws = wb[name]
				if ws.cell(row=1, column=1).value == "Документ":
					sheets_to_process.append(name)

		if not sheets_to_process:
			print("Нет подходящих листов для обработки")
			return 0

		output_dir = self.project.project_BoQs_path
		output_dir.mkdir(parents=True, exist_ok=True)

		processed = 0
		for sheet_name in sheets_to_process:
			ws = wb[sheet_name]

			# --- Чтение метаданных из шапки (строки 1-16) ---
			meta = {}
			for row in ws.iter_rows(min_row=1, max_row=16, max_col=14, values_only=True):
				if not row[0]:
					continue
				key = str(row[0]).strip()
				if key == "Наименование стройки":
					meta['site'] = row[3]
				elif key == "Наименование объекта капитального строительства":
					meta['object'] = row[3]
				elif key == "Ведомость объемов работ №":
					meta['num'] = row[3]
				elif key == "Основание((наименование раздела (подраздела) ПД))":
					meta['reason'] = row[3]
				elif key == "Дата составления":
					meta['date'] = row[3]
				elif key == "Составил ФИО":
					meta['composer_name'] = row[3]
				elif key == "Составил должность":
					meta['composer_pos'] = row[3]
				elif key == "Проверил ФИО":
					meta['verifier_name'] = row[3]
				elif key == "Проверил должность":
					meta['verifier_pos'] = row[3]

			# Преобразование даты
			date_obj = None
			if meta.get('date'):
				try:
					if isinstance(meta['date'], datetime):
						date_obj = meta['date']
					else:
						date_str = str(meta['date']).split()[0]
						date_obj = datetime.strptime(date_str, "%Y-%m-%d")
				except Exception:
					date_obj = datetime.now()
			else:
				date_obj = datetime.now()
			date_str = date_obj.strftime("%d.%m.%Y")

			# Формирование metadata для JSON
			metadata_json = {
				"ObjectName": clean_string(meta.get('object')),
				"Num": clean_string(meta.get('num')),
				"local_estimate": "",
				"Date": date_str,
				"Signatures": {
					"Composer": clean_string(meta.get('composer_name')),
					"Composer_Position": clean_string(meta.get('composer_pos'))
				},
				"Status_Done": False,
				"log_list": [],
				"note": ""
			}

			# --- Парсинг разделов, работ и ресурсов ---
			sections = []
			current_section = None
			current_work = None	  # последняя созданная работа (для добавления ресурсов)
			current_element = None   # последний созданный элемент (работа или ресурс) для ссылок

			start_row = 17
			for row in ws.iter_rows(min_row=start_row, values_only=False):
				# Чтение всех необходимых колонок (0-based индексы)
				a = row[0].value   # № п.п.
				b = row[1].value   # Наименование
				c = row[2].value   # Ед. изм.
				e = row[4].value   # Формула расчёта
				f = row[5].value   # Описание ссылки (не используется в JSON)
				g = row[6].value   # Путь/имя файла (не используется)
				h = row[7].value   # Номера страниц -> user_pages
				i = row[8].value   # Тег -> tag
				j = row[9].value   # Комментарий
				l_val = row[11].value  # Тип
				m_val = row[12].value if len(row) > 12 else None  # book_num (колонка M)

				# Пропуск полностью пустых строк
				if all(v is None for v in [a, b, c, e, f, g, h, i, j, l_val, m_val]):
					continue

				# Проверка на начало раздела
				section_name = detect_section(row)
				if section_name:
					current_section = {
						"name": section_name,
						"works": []
					}
					sections.append(current_section)
					current_work = None
					current_element = None
					continue

				# Определение номера работы
				work_num = parse_work_number(a)

				# Строка с номером (новая работа или ресурс)
				if work_num is not None:
					typ = clean_string(l_val).lower() if l_val else "работа"
					raw_name = clean_string(b)
					raw_unit = find_unit_key(clean_string(c))
					raw_quantity_formula = clean_string(e) if e else ""
					raw_comment = clean_string(j)

					if typ == "работа":
						# Создаём новую работу
						work = {
							"raw_name": raw_name,
							"raw_unit": raw_unit,
							"raw_quantity_formula": raw_quantity_formula,
							"raw_comment": raw_comment,
							"type": typ,
							"links": [],
							"resources": []
						}
						if current_section is None:
							# Создаём раздел по умолчанию, если его ещё нет
							current_section = {"name": "Основные работы", "works": []}
							sections.append(current_section)
						current_section["works"].append(work)
						current_work = work
						current_element = work
					else:
						# Это ресурс (материал, оборудование и т.д.)
						if current_work is not None:
							resource = {
								"raw_name": raw_name,
								"raw_unit": raw_unit,
								"raw_quantity_formula": raw_quantity_formula,
								"raw_comment": raw_comment,
								"type": typ,
								"links": []
							}
							current_work["resources"].append(resource)
							current_element = resource
						else:
							# Ресурс без родительской работы – игнорируем или можно создать отдельную работу
							print(f"Предупреждение: ресурс '{raw_name}' без родительской работы, пропущен")
							current_element = None
				else:
					# Строка без номера – добавляем ссылку к текущему элементу (работе или ресурсу)
					if current_element is not None and (g or h):
						book_num = m_val if m_val is not None else None
						tag = i if i is not None else None
						user_pages = normalize_page_numbers(h)
						link = {
							"book_num": book_num,
							"tag": tag,
							"user_pages": user_pages
						}
						current_element["links"].append(link)

			# --- Постобработка: можно добавить фиктивные ссылки, если требуется (по желанию) ---
			# В текущей реализации оставляем пустые links, как в примере JSON

			# --- Формирование итогового JSON ---
			output_data = {
				"metadata": metadata_json,
				"data": {
					"sections": sections,
					"archive": []
				}
			}

			# --- Сохранение в файл ---
			#safe_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
			output_path = output_dir / f"{filepath.stem}.json"
			with open(output_path, "w", encoding="utf-8") as f:
				json.dump(output_data, f, indent=4, ensure_ascii=False)

			print(f"Конвертирован лист '{sheet_name}' -> {output_path}")
			processed += 1

		wb.close()
		return processed
	
	def export_to_pdf(self, manager: BoQ_manager, mode = 0) -> bool:
		"""
		Экспорт ведомости в PDF.
		:param manager: объект BoQ_manager с данными
		:param mode: форма верхней части
		:return: True при успехе
		"""
		try:
			generator = PDFGenerator(manager, self.project, mode)
			folder = manager.file.parent.parent / 'PDF'
			folder.mkdir(parents=True, exist_ok=True)
			output = folder / manager.file.stem
			generator.generate(output.with_suffix('.pdf'))
			return True
		except Exception as e:
			print(f"Ошибка генерации PDF: {e}")
			return False



class BookmarkParagraph(Paragraph):
	def __init__(self, text, style, bookmark_name, bookmark_title):
		super().__init__(text, style)
		self.bookmark_name = bookmark_name
		self.bookmark_title = bookmark_title
		self.already_bookmarked = False

	def draw(self):
		# Добавляем закладку при первом рисовании
		if not self.already_bookmarked:
			canvas = self.canv
			# Закладка на текущую позицию (по горизонтали не сдвигаем)
			canvas.bookmarkHorizontal(self.bookmark_name, 0, 0)
			# Добавляем запись в оглавление (закладку)
			canvas.addOutlineEntry(
				self.bookmark_title,
				self.bookmark_name,
				level=0,	# корневой уровень
				closed=False
			)
			self.already_bookmarked = True
		super().draw()


class PDFGenerator:
	"""Генератор PDF ведомости объемов работ."""

	# Ширины столбцов (в мм) из спецификации
	COL_WIDTHS_MM = [14, 75, 18, 25, 45, 55, 45]

	def __init__(self, manager, project, mode):
		self.manager: BoQ_manager = manager
		self.project: Project = project
		self.mode: int = mode
		# Регистрация шрифта Arial (если доступен в системе, иначе используется стандартный Helvetica)
		try:
			pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
			self.font_name = 'Arial'

			pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
			self.bold_font_name='Arial-Bold'

			pdfmetrics.registerFont(TTFont('Arial-Italic', 'ariali.ttf'))
			self.italic_font_name = 'Arial-Italic'

		except:
			self.font_name = 'Helvetica'

	def mm_to_pts(self, mm_value: float) -> float:
		"""Перевод миллиметров в пункты (1 мм = 2.83465 pt)."""
		return mm_value * 2.83465

	def _build_styles(self):
		"""Создание стилей для документа."""
		styles = getSampleStyleSheet()
		# Стиль для обычного текста (11 pt, Arial)
		normal_style = ParagraphStyle(
			'CustomNormal',
			parent=styles['Normal'],
			fontName=self.font_name,
			fontSize=11,
			leading=13,
			alignment=TA_LEFT
		)
		# Стиль для выравнивания по центру
		center_style = ParagraphStyle(
			'CenterStyle',
			parent=normal_style,
			alignment=TA_CENTER
		)
		# Стиль для жирного текста 12 pt (заголовок ведомости)
		bold_12_center = ParagraphStyle(
			'Bold12Center',
			parent=center_style,
			fontSize=12,
			fontName=self.bold_font_name,
			leading=14
		)
		# Стиль для курсивных подписей
		italic_style = ParagraphStyle(
			'ItalicStyle',
			parent=normal_style,
			fontName=self.italic_font_name,
			fontSize=11,
			alignment=TA_LEFT
		)
		# Стиль для курсивных подписей по центру
		italic_center_style = ParagraphStyle(
			'ItalicStyle',
			parent=normal_style,
			fontName=self.italic_font_name,
			fontSize=11,
			alignment=TA_CENTER
		)

		# Стиль для заголовка таблицы (жирный, по центру)
		table_header_style = ParagraphStyle(
			'TableHeader',
			parent=normal_style,
			fontName=self.bold_font_name,
			fontSize=10,
			alignment=TA_CENTER,
			textColor=colors.black
		)
		# Стиль для ячеек таблицы (центр)
		table_cell_center = ParagraphStyle(
			'TableCellCenter',
			parent=normal_style,
			alignment=TA_CENTER
		)
		# Стиль для ячеек таблицы (лево)
		table_cell_left = ParagraphStyle(
			'TableCellLeft',
			parent=normal_style,
			fontSize=10,
			alignment=TA_LEFT
		)
		# Стиль для строки раздела (жирный, лево)
		section_style = ParagraphStyle(
			'SectionStyle',
			parent=normal_style,
			fontName=self.bold_font_name,
			fontSize=10,
			alignment=TA_LEFT
		)
		return {
			'normal': normal_style,
			'center': center_style,
			'bold_12_center': bold_12_center,
			'italic': italic_style,
			'italic_center': italic_center_style,
			'table_header': table_header_style,
			'table_cell_center': table_cell_center,
			'table_cell_left': table_cell_left,
			'section': section_style
		}

	def _prepare_table_data(self, styles):
		"""
		Подготовка данных для таблицы в виде списка строк.
		Возвращает:
			- table_data: list[list[Paragraph]] - строки таблицы
			- row_keep_together: list[int] - индексы строк, которые нельзя разрывать (разделы)
		"""
		table_data = []
		row_keep_together = []

		# Заголовки столбцов
		headers = [
			"№<br/>п/п",
			"Наименование работ, ресурсов,<br/>затрат по проекту",
			"Ед.<br/>изм.",
			"Объем работ/<br/>Количество",
			"Формула расчета объемов работ и расхода материалов, потребности ресурсов",
			"Ссылка на чертежи, спецификации в проектной документации",
			"Дополнительная информация (комментарий)"
		]
		header_paragraphs = [Paragraph(h, styles['table_header']) for h in headers]
		table_data.append(header_paragraphs)

		# Проход по разделам и работам
		for section in self.manager.sections:
			# Строка раздела (объединение всех ячеек)
			section: Section
			section_text = f"<b>{section.name}</b>" if section.name else "Раздел"
			# Уникальное имя закладки: используем номер строки или имя раздела + номер секции
			bookmark_name = f"sec_{len(table_data)}"
			bookmark_title = section.name if section.name else "Раздел"
			section_para = BookmarkParagraph(section_text, styles['section'], bookmark_name, bookmark_title)
			section_row = [section_para] + [Paragraph("", styles['normal']) for _ in range(6)]
			table_data.append(section_row)
			row_keep_together.append(len(table_data)-1)  # эту строку нельзя разрывать

			# Работы и их ресурсы
			for work in section.works:
				# Добавляем строку работы
				work_row = self._make_work_row(work, styles)
				table_data.append(work_row)
				# Для ресурсов добавляем отступ в наименовании
				for resource in getattr(work, 'resources', []):
					resource_row = self._make_resource_row(resource, styles)
					table_data.append(resource_row)

		return table_data, row_keep_together
	
	def _make_work_row(self, work: Work, styles):
		"""Создает строку таблицы для работы."""
		# Номер работы
		num = Paragraph(str(work.num) if work.num else "", styles['table_cell_center'])
		# Наименование
		name = Paragraph(work.name or "", styles['table_cell_left'])
		# Ед. изм.
		unit = Paragraph(work.unit.replace(' ', '<br/>') or "", styles['table_cell_center'])
		# Объем
		quantity = Paragraph(str(work.quantity) if work.quantity is not None else "", styles['table_cell_center'])
		# Формула
		formula = Paragraph(work.quantity_formula or "", styles['table_cell_center'])
		# Ссылка
		link_text = work.planned_links
		link_para = Paragraph(link_text, styles['table_cell_left'])
		# Комментарий
		comment = Paragraph(work.comment or "", styles['table_cell_left'])
		return [num, name, unit, quantity, formula, link_para, comment]

	def _make_resource_row(self, resource: Resource, styles):
		"""Создает строку таблицы для ресурса с отступом в наименовании."""
		num = Paragraph(str(resource.num) if resource.num else "", styles['table_cell_center'])
	
		indented_name = resource.name if resource.name else ""
		name = Paragraph(indented_name, styles['table_cell_left'])
		unit = Paragraph(resource.unit or "", styles['table_cell_center'])
		quantity = Paragraph(str(resource.quantity) if resource.quantity is not None else "", styles['table_cell_center'])
		formula = Paragraph(resource.quantity_formula or "", styles['table_cell_center'])
		link_text = resource.planned_links
		link_para = Paragraph(link_text, styles['table_cell_left'])
		comment = Paragraph(resource.comment or "", styles['table_cell_left'])
		return [num, name, unit, quantity, formula, link_para, comment]


	def _build_header_flowables(self, styles):
		"""Создаёт flowables для верхней части документа (пункты 1-13)."""
		flowables = []

		# Пункт 1: Наименование объекта (данные)
		site_name = self.project.construction_site
		flowables.append(Paragraph(site_name, styles['center']))

		# Пункт 2: подчёркивающая черта на всю ширину
		flowables.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceBefore=0, spaceAfter=0))

		# Пункт 3: (расшифровка)
		flowables.append(Paragraph(f"(наименование стройки)", styles['italic_center']))

		# Интервал (пустая строка)
		flowables.append(Spacer(1, 5*mm))

		# Пункт 5: Наименование ведомости (данные – можно взять из project или задать константу)
		doc_name = self.manager.object_name or ''
		flowables.append(Paragraph(doc_name, styles['center']))

		# Пункт 6: подчёркивающая черта
		flowables.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceBefore=0, spaceAfter=0))

		# Пункт 7: (расшифровка)
		flowables.append(Paragraph(f"(наименование объекта капитального строительства)", styles['italic_center']))

		# Двойной интервал (две пустые строки)
		flowables.append(Spacer(1, 10*mm))

		# Пункт 9: "Ведомость объемов работ № ВО-" с данными (жирный, 12 пт)
		num_str = f"Ведомость объемов работ № {self.manager.num or ''}"
		flowables.append(Paragraph(num_str, styles['bold_12_center']))

		# Интервал
		flowables.append(Spacer(1, 10*mm))

		# Пункт 11: "Основание: ..." (выравнивание слева)
		reason_label = Paragraph("Основание:", styles['normal'])
		reason_value = Paragraph(self.manager.reason or '', styles['normal'])
		reason_table = Table([[reason_label, reason_value]], colWidths=[self.mm_to_pts(40), None])
		reason_table.setStyle(TableStyle([
			('VALIGN', (0,0), (-1,-1), 'TOP'),
			('ALIGN', (0,0), (0,0), 'RIGHT'),	# метка выровнена вправо
			('ALIGN', (1,0), (1,0), 'LEFT'),	 # данные – влево
			('LEFTPADDING', (0,0), (0,0), 0),
			('RIGHTPADDING', (1,0), (1,0), 0),
		]))
		flowables.append(reason_table)

		# Пункт 12: "Дата составления: ..."
		date_label = Paragraph("Дата составления:", styles['normal'])
		date_value = Paragraph(self.manager.date or self.project.now.strftime("%d.%m.%Y"), styles['normal'])
		date_table = Table([[date_label, date_value]], colWidths=[self.mm_to_pts(40), None])
		date_table.setStyle(TableStyle([
			('VALIGN', (0,0), (-1,-1), 'TOP'),
			('ALIGN', (0,0), (0,0), 'RIGHT'),
			('ALIGN', (1,0), (1,0), 'LEFT'),
			('LEFTPADDING', (0,0), (0,0), 0),
			('RIGHTPADDING', (1,0), (1,0), 0),
		]))
		flowables.append(date_table)

		# Двойной интервал перед таблицей
		flowables.append(Spacer(1, 10*mm))

		return flowables

	def _build_signatures_flowables(self, styles):
		"""Создаёт flowables для подписей после таблицы."""
		flowables = []
		line = '_____________________'
		transript = "[должность, подпись (инициалы, фамилия)]"

		flowables.append(Spacer(1, 5*mm))

		# Составил
		composer = self.manager.signatures.get('Composer', '')
		composer_pos = self.manager.signatures.get('Composer_Position', '')
		flowables.append(Paragraph(f"Составил: {composer_pos}, {composer}		{line}", styles['normal']))
	
		# Расшифровка курсивом
		flowables.append(Paragraph(transript, styles['italic']))
		flowables.append(Spacer(1, 5*mm))

		# Проверил
		if isinstance(self.manager.verifier, dict):
			verifier_data = self.manager.verifier
		else:
			verifier_data = self.project.verifier
		verifier = verifier_data.get('Name', '')
		verifier_pos = verifier_data.get('Position', '')
		flowables.append(Paragraph(f"Проверил: {verifier_pos}, {verifier}		{line}", styles['normal']))

		flowables.append(Paragraph(transript, styles['italic']))

		return flowables

	def _build_compact_header_and_signatures(self, styles):
		"""Альтернативное оформление шапки и подписей в виде единой таблицы."""
		flowables = []
		
		# Ширина первой колонки (под метки) – 60 мм, вторая – всё остальное
		col_widths = [self.mm_to_pts(60), None]
		
		# --- Строка 1: Документ / Ведомость объемов работ ---

		if isinstance(self.manager.verifier, dict):
			verifier_data = self.manager.verifier
		else:
			verifier_data = self.project.verifier

		data_rows = [
			["Документ", "Ведомость объемов работ"],
			["Версия", "3_01"],
			["", ""],   # пустая разделительная строка
			["Наименование стройки", self.project.construction_site],
			["Наименование объекта капитального строительства", self.manager.object_name or ''],
			["Ведомость объемов работ №", self.manager.num or ''],
			["Основание", self.manager.reason.replace('\n', '<br/>') or ''],
			["Дата составления", self.manager.date or self.project.now.strftime("%d.%m.%Y")],
			["", ""],   # пустая строка перед подписями
			["Составил ФИО", self.manager.signatures.get('Composer', '')],
			["Составил должность", self.manager.signatures.get('Composer_Position', '')],
			["Проверил ФИО", verifier_data.get('Name', '')],
			["Проверил должность", verifier_data.get('Position', '')],
		]
		
		# Преобразуем в Paragraph для корректного переноса
		table_data = []
		for label, value in data_rows:
			label_para = Paragraph(label, styles['normal'])
			value_para = Paragraph(value, styles['normal'])
			table_data.append([label_para, value_para])
		
		# Создаём таблицу
		table = Table(table_data, colWidths=col_widths)
		table.setStyle(TableStyle([
			('ALIGN', (0,0), (0,-1), 'RIGHT'),	  # метки выравниваем вправо
			('ALIGN', (1,0), (1,-1), 'LEFT'),	   # значения – влево
			('VALIGN', (0,0), (-1,-1), 'TOP'),
			('LEFTPADDING', (0,0), (0,-1), 0),	  # убираем лишние отступы слева у меток
			('RIGHTPADDING', (1,0), (1,-1), 0),
			('TOPPADDING', (0,0), (-1,-1), 2),
			('BOTTOMPADDING', (0,0), (-1,-1), 2),
			('FONTNAME', (0,0), (-1,-1), self.font_name),
			('FONTSIZE', (0,0), (-1,-1), 11),
		]))
		
		# Добавляем горизонтальную черту после строки "Версия" (опционально)
		# Можно просто добавить Spacer или пустую строку – уже есть
		
		flowables.append(table)
		flowables.append(Spacer(1, 10*mm))
		return flowables

	def add_page_number(self, canvas, doc):
		canvas.saveState()
		page_num = canvas.getPageNumber()
		canvas.setFont(self.font_name, 9)
		canvas.drawRightString(doc.width + doc.leftMargin - 10, doc.bottomMargin - 5, f"Лист {page_num}")
		canvas.restoreState()

	def generate(self, output_path: Path):
		"""Основной метод генерации PDF."""
		styles = self._build_styles()
		# Подготовка данных таблицы
		table_data, keep_rows = self._prepare_table_data(styles)

		# Преобразование ширины столбцов в пункты
		col_widths = [self.mm_to_pts(w) for w in self.COL_WIDTHS_MM]

		# Создание таблицы
		table = Table(table_data, colWidths=col_widths, repeatRows=1)
		# Стили таблицы
		tbl_style = TableStyle([
			('FONTNAME', (0,0), (-1,-1), self.font_name),
			('FONTSIZE', (0,0), (-1,-1), 11),
			('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
			('GRID', (0,0), (-1,-1), 0.5, colors.black),
			('BOX', (0,0), (-1,-1), 1, colors.black),
		])

		# Выравнивание для конкретных столбцов
		# Столбцы с индексами 0,2,3,4 – по центру (все остальные по левому краю)
		# Выравнивание для каждого столбца (0-6)
		col_alignments = ['CENTER', 'LEFT', 'CENTER', 'CENTER', 'CENTER', 'LEFT', 'LEFT']
		for row_idx in range(len(table_data)):
			for col_idx, align in enumerate(col_alignments):
				tbl_style.add('ALIGN', (col_idx, row_idx), (col_idx, row_idx), align)

		# Объединение ячеек для строк разделов (индексы из keep_rows)
		for row_idx in keep_rows:
			tbl_style.add('SPAN', (0, row_idx), (-1, row_idx))
			# Для объединённой строки задаём выравнивание по левому краю (можно и по центру, но по ТЗ не уточнено)
			tbl_style.add('ALIGN', (0, row_idx), (-1, row_idx), 'LEFT')

		# Установка стиля для строк-разделов – жирный шрифт уже задан в Paragraph
		table.setStyle(tbl_style)

		# Сборка документа
		doc = SimpleDocTemplate(
			str(output_path),
			pagesize=landscape(A4),
			leftMargin=self.mm_to_pts(10),   	# 10 мм
			rightMargin=self.mm_to_pts(10),  	# 10 мм
			topMargin=self.mm_to_pts(20),		# 20 мм
			bottomMargin=self.mm_to_pts(10)  	# 10 мм
		)

		# Содержимое документа: заголовки, таблица, подписи
		story = []
		if self.mode == 0:
			story.extend(self._build_header_flowables(styles))
		else:
			story.extend(self._build_compact_header_and_signatures(styles))

		# Обёртывание таблицы в KeepTogether для того, чтобы строки-разделы не разрывались?
		# Просто добавляем таблицу – splitByRow=True по умолчанию, и если строка не влезает, она переносится.
		# Для дополнительной гарантии оборачиваем каждую "чувствительную" строку? Но проще положиться на Table.
		story.append(table)

		if self.mode == 0:
			story.extend(self._build_signatures_flowables(styles))

		# Построение документа
		doc.build(story, onFirstPage=self.add_page_number, onLaterPages=self.add_page_number)


class ShemaError(Exception):
	def __init__(self, filename, type = None, vers = None):
		super().__init__()
		self.filename = filename
		self.type = type
		self.version = vers
	
	def info(self):
		if not self.type or not self.version:
			return f'Файл {self.filename} не соответствует схеме'
		else:
			return f'Файл {self.filename} не соответствует схеме. [Тип схемы: {self.type}, Версия: {self.version}]'
