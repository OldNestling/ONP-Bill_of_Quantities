# Copyright © 2026 OldNestling
# License: GPLv3 (GNU General Public License Version 3)

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

from pathlib import Path
from openpyxl import load_workbook

from PyQt6.QtWidgets import (
	QApplication, QWidget, QLabel, QPushButton, QFileDialog, QDialog, 
	QDialogButtonBox, QVBoxLayout, QHBoxLayout, QGroupBox, QLineEdit,
	QPlainTextEdit, QMessageBox, QListWidget, QAbstractItemView, QTableWidget, 
	QTableWidgetItem, QHeaderView, QComboBox, QCompleter, QTabWidget, QFrame, 
	QProgressDialog, QInputDialog, QTextBrowser, QMenu)
from PyQt6.QtCore import Qt, QSize, QTimer
from PyQt6.QtGui import QColor, QPixmap

from Core.Project import Project, File_BoQ, FileIsBusy
from Core.Utilities import text_after, text_before
from Core.Convertor_BoQ import Convertor, ShemaError
from Templates.About import (
	ABOUT, COPYRIGHT, GITHUB, PROGRAM_NAME, PROGRAM_VERSION, PARTICIPANTS
)
from .ui_utilities import MultilineTextDelegate, create_separator, Requestion
from .tabs.settings_tab import SettingsTab 
from .tabs.soils_tab import Soils_Tab 
from .tabs.source_tab import Sources_Tab
from .tabs.machinery_tab import Machinery_Tab
from .tabs.documentation_tab import Documentation_Tab
from .tabs.work_tab import BoQ_Tab
from .tabs.user_libs_tab import User_Libs_Tab
from .icons import Icons


class Project_Window(QWidget):
	""" Контейнер с вкладками """
	def __init__(self):
		super().__init__()
		self.project_directory = None
		self.project: Project = None
		self.main_tab = None			# ссылка на главную вкладку
		self.other_tabs = []			# список остальных вкладок
		self.initializeUI()

	def initializeUI(self):
		self.setWindowTitle('Система работы с ВОР')

		# Создаём TabWidget
		self.tab_widget = QTabWidget()
		self.tab_widget.currentChanged.connect(self.tab_changed)

		# Создаём главную вкладку
		self.main_tab = MainProjectTab(self)   # передаём ссылку на главное окно
		self.tab_widget.addTab(self.main_tab, "Проект")

		# Создаём остальные вкладки (заглушки)
		self.create_other_tabs()

		# Основной layout
		main_layout = QVBoxLayout()
		main_layout.addWidget(self.tab_widget)
		self.setLayout(main_layout)

		self.show()
		self.resize(900, 600)
		self.center_window()

	def create_other_tabs(self):
		"""Создаёт дополнительные вкладки и блокирует их."""
		
		self.BoQs_tab = BoQ_Tab(project=self.project, main_window=self)
		self.tab_widget.addTab(self.BoQs_tab, "Ведомости")
		self.other_tabs.append(self.BoQs_tab)

		self.settings_tab = SettingsTab(project=self.project, main_window=self)
		self.tab_widget.addTab(self.settings_tab, "Настройки")
		self.other_tabs.append(self.settings_tab)
		
		soils_tab = Soils_Tab(project=self.project)
		self.tab_widget.addTab(soils_tab, "Грунты")
		self.other_tabs.append(soils_tab)

		transportation_tab = Sources_Tab(project=self.project)
		self.tab_widget.addTab(transportation_tab, "Источники и транспортировка")
		self.other_tabs.append(transportation_tab)

		userlib_tab = User_Libs_Tab(project=self.project)
		self.tab_widget.addTab(userlib_tab, "Пользовательские библиотеки")
		self.other_tabs.append(userlib_tab)

		machines_tab = Machinery_Tab(project=self.project)
		self.tab_widget.addTab(machines_tab, "Механизация")
		self.other_tabs.append(machines_tab)

		content_tab = Documentation_Tab(project=self.project)
		self.tab_widget.addTab(content_tab, "Содержание ПД")
		self.other_tabs.append(content_tab)

		#typical_products_tab = QWidget()	TODO
		#typical_products_tab.setLayout(QVBoxLayout())
		#typical_products_tab.layout().addWidget(QLabel("Область настройки типовых изделий.\nБудет реализовано в поздних версиях"))
		#self.tab_widget.addTab(typical_products_tab, "Типовые изделия")
		#self.other_tabs.append(typical_products_tab)

		#typical_work_tab = QWidget()	TODO
		#typical_work_tab.setLayout(QVBoxLayout())
		#typical_work_tab.layout().addWidget(QLabel("Область настройки типовых работ.\nБудет реализовано в поздних версиях"))
		#self.tab_widget.addTab(typical_work_tab, "Типовые работы")
		#self.other_tabs.append(typical_work_tab)


		# Блокируем дополнительные вкладки
		self.set_other_tabs_enabled(False)

	def set_other_tabs_enabled(self, enabled):
		"""Включает/отключает дополнительные вкладки (кроме первой)."""
		for i in range(1, self.tab_widget.count()):
			self.tab_widget.setTabEnabled(i, enabled)

	def update_all_tabs_project(self):
		"""Обновляет ссылку на проект во всех вкладках (кроме главной)"""
		for tab in self.other_tabs:
			if hasattr(tab, 'set_project'):
				tab.set_project(self.project)
	
	def tab_changed(self, index):
		"""Вызывается при переключении вкладки"""
		current_tab = self.tab_widget.widget(index)
		if hasattr(current_tab, 'tab_selected'):
			current_tab.tab_selected()
	
	def notify_settings_changed(self):
		"""Вызывается при изменении настроек проекта (например, переключении режимов)."""
		for tab in self.other_tabs:
				if hasattr(tab, 'settings_changed'):
					tab.settings_changed()
	def closeEvent(self, event):
		# Если есть вкладка "Разделы" и в ней открыты файлы
		if hasattr(self, 'BoQs_tab') and self.BoQs_tab is not None:
			if self.BoQs_tab.boq_views_container.count() > 0:
				if self.BoQs_tab.has_unsaved_changes():
					reply = Requestion.ask(
						self,
						'Несохранённые изменения',
						'В открытых разделах есть несохранённые изменения. Сохранить перед выходом?',
					)
					if reply == QMessageBox.StandardButton.Cancel:
						event.ignore()
						return
					elif reply == QMessageBox.StandardButton.Yes:
						success = self.BoQs_tab.close_all_tabs(ask_for_save=True, parent_dialog=self)
						if not success:
							event.ignore()
							return
					else:  # No – не сохранять
						self.BoQs_tab.force_close_all()
				else:
					# Нет изменений – просто закрываем вкладки
					self.BoQs_tab.force_close_all()

		# Сохраняем настройки проекта
		if self.project:
			self.project.saving_settings()

		event.accept()

	def flash_tab(self, index, color=Qt.GlobalColor.green, duration=500):
		"""Мигает вкладкой с заданным цветом на короткое время."""
		tab_bar = self.tab_widget.tabBar()
		if tab_bar is None:
			return
		original_color = tab_bar.tabTextColor(index)
		tab_bar.setTabTextColor(index, QColor(color))
		# Возвращаем исходный цвет через duration мс
		QTimer.singleShot(
			duration, 
			lambda: tab_bar.setTabTextColor(
				index, original_color if original_color is not None else QColor()
			)
		)

	# ============================= Вспомогательные функции =============================
	def get_project_dir(self):
		if self.project_directory is None:
			return 'Директория ВОР проекта не установлена'
		else:
			return self.project_directory

	def set_project_dir(self, new_dir: Path = None):
		""" Задаёт рабочию директорию для проекта
		:new_dir: путь к новой ревизии, получаемый при её создании """
		# Если есть открытые вкладки с изменениями — обрабатываем как было
		if hasattr(self, 'BoQs_tab') and self.BoQs_tab is not None:
			if self.BoQs_tab.boq_views_container.count() > 0:
				if self.BoQs_tab.has_unsaved_changes():
					reply = Requestion.ask(
						self,
						"Несохранённые изменения",
						"В открытых разделах есть несохранённые изменения. Сохранить перед выходом?"
					)
					if reply == QMessageBox.StandardButton.Cancel:
						return
					elif reply == QMessageBox.StandardButton.Yes:
						success = self.BoQs_tab.close_all_tabs(ask_for_save=True, parent_dialog=self)
						if not success:
							return
					else:  # No – не сохранять
						self.BoQs_tab.force_close_all()
				else:
					self.BoQs_tab.force_close_all()
		# Если передан путь новой ревизии – используем его напрямую
		if new_dir is not None:
			selected_dir = str(new_dir)
			Project.save_recent_dir(selected_dir)
		else:
			# Создаём диалог выбора директории
			recent_dirs = Project.load_recent_dirs()
			dialog = QDialog(self)
			dialog.setWindowTitle("Выбор директории проекта")
			dialog.setMinimumWidth(600)
			dialog.setModal(True)
			layout = QVBoxLayout(dialog)

			if recent_dirs:
				recent_dirs.reverse()
				layout.addWidget(QLabel("<p>Недавние папки <i>(выбор двойным нажатием)</i>:</p>"))
				list_widget = QListWidget()
				for path in recent_dirs:
					list_widget.addItem(path)
				list_widget.itemDoubleClicked.connect(lambda: dialog.accept())
				layout.addWidget(list_widget)

			btn_browse = QPushButton("Выбрать другую папку...")
			layout.addWidget(btn_browse)


			# Обработка выбора
			selected_dir = None
			def on_browse():
				nonlocal selected_dir
				dir_path = QFileDialog.getExistingDirectory(
					self, "Выберите папку проекта", "",
					QFileDialog.Option.ShowDirsOnly
				)
				if dir_path:
					selected_dir = dir_path
					dialog.accept()
			btn_browse.clicked.connect(on_browse)

			# Если есть список последних, при двойном щелчке выбираем из него
			if recent_dirs:
				def on_item_selected():
					nonlocal selected_dir
					current = list_widget.currentItem()
					if current:
						selected_dir = current.text()
						dialog.accept()
				list_widget.itemDoubleClicked.connect(on_item_selected)

			if dialog.exec() != QDialog.DialogCode.Accepted or not selected_dir:
				return

			# Сохраняем выбранную директорию в историю
			Project.save_recent_dir(selected_dir)

		# Создаём или пересоздаём проект
		self.project_directory = selected_dir
		self.project = Project(base_dir=selected_dir, sourсes_folder=r'Data')
		# Разблокируем остальные вкладки
		self.set_other_tabs_enabled(True)
		# Уведомляем главную вкладку
		self.main_tab.on_project_dir_changed()
		self.update_all_tabs_project()

	def save_project_data(self):
		"""Сохраняет настройки проекта, если проект выбран."""
		if self.project is not None:
			self.project.saving_settings()
			QMessageBox.information(self, "Сохранение", "Настройки проекта сохранены.")
		else:
			print("Невозможно сохранить: проект не выбран")

	def center_window(self):
		screen = QApplication.primaryScreen().availableGeometry()
		window_size = self.geometry()
		x = (screen.width() - window_size.width()) // 2
		y = (screen.height() - window_size.height()) // 2
		self.move(x, y)

	def open_project_folder(self):
		""" Открывает текущую папку проекта """
		if self.project is None:
			QMessageBox.warning(self, "Открыть папку проекта", "Проект не загружен.")
			return
		
		self.project.open_project_folder()

	def new_refinement(self):
		"""Создаёт новую ревизию текущего проекта и переключается на неё."""
		if self.project is None:
			QMessageBox.warning(self, "Новая ревизия", "Проект не загружен.")
			return

		# Спрашиваем имя новой папки (будет создана рядом с текущей)
		parent_dir = self.project.base_dir.parent
		default_name = self.project.base_dir.stem + "_копия"
		name, ok = QInputDialog.getText(
			self,
			"Новая ревизия",
			f"Введите имя папки для новой ревизии:\n(будет создана в {parent_dir})",
			text=default_name
		)
		if not ok or not name.strip():
			return

		new_path = self.project.create_new_refinement(name.strip())
		if new_path is None:
			QMessageBox.warning(self, "Новая ревизия", f"Папка '{name}' уже существует.")
			return

		# Переключаем проект на новую ревизию
		self.set_project_dir(new_dir=new_path)

	def get_help(self):
		"""Выводит структурированную справку о программе."""
		dialog = QDialog(self)
		dialog.setWindowTitle("Справка")
		dialog.setModal(True)
		dialog.setMinimumWidth(550)
		dialog.setMinimumHeight(400)

		# Установка иконки окна (если есть)
		dialog.setWindowIcon(Icons.recolor_icon(Icons.help, QColor("#000000")))

		main_layout = QVBoxLayout(dialog)

		# Группа "О программе"
		about_group = QGroupBox("О программе")
		about_layout = QVBoxLayout()

		# Название и версия
		title_label = QLabel(f"<h2>{PROGRAM_NAME}</h2>")
		title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
		about_layout.addWidget(title_label)

		version_label = QLabel(f'<b>Версия:</b> {PROGRAM_VERSION}')
		version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
		about_layout.addWidget(version_label)

		copyright_label = QLabel(f'<i>{COPYRIGHT} (GPLv3)</i>')
		copyright_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
		about_layout.addWidget(copyright_label)

		about_layout.addWidget(create_separator(QFrame.Shape.HLine))
	
		# Описание программы
		about_text = QTextBrowser()
		about_text.setOpenExternalLinks(True)
		about_text.setHtml(f"<p>{ABOUT.replace(chr(10), '<br>')}</p>")
		about_text.setMaximumHeight(120)
		about_text.setMinimumHeight(60)
		about_layout.addWidget(about_text)
		about_layout.addStretch()

		about_group.setLayout(about_layout)

		# Группа "Справка и контакты"
		info_group = QGroupBox("Справка и контакты")
		info_layout = QVBoxLayout()

		# Ссылка на учебное пособие
		tutorial_label = QLabel()
		tutorial_label.setOpenExternalLinks(True)
		tutorial_label.setText(f'<b>Справочная информация: </b> <a href="{GITHUB}"style="color: #3B82F6; text-decoration: none;">{GITHUB}</a>')
		tutorial_label.setWordWrap(True)
		info_layout.addWidget(tutorial_label)

		# Контакты
		contact_label = QLabel(
			f'<b>Связь с разработчиком: </b>'
			f'<a href="mailto:oldnestling@yandex.ru" style="color: #3B82F6; text-decoration: none;">oldnestling@yandex.ru</a>'
			f'<br><b>Счёт для благодарности: </b>'
			f'<a href="https://www.tinkoff.ru/rm/r_aOqwUwfbRm.lZNlrKjSLm/oMI8k46003" style="color: #3B82F6; text-decoration: none;">4081 7810 6000 9641 9718</a>'
		)
		contact_label.setOpenExternalLinks(True)
		# contact_label.setTextFormat(Qt.TextFormat.RichText)  # явно указываем rich‑text
		# contact_label.setStyleSheet("QLabel { background-color: #e1e1e1; color: #000000; } a { color: #3B82F6; }")
		contact_label.setWordWrap(True)

		info_layout.addWidget(contact_label)

		participants = QLabel(PARTICIPANTS
		)
		participants.setWordWrap(True)

		info_layout.addWidget(participants)

		# Дополнительная информация (можно добавить благодарности, лицензию и т.п.)
		info_layout.addWidget(create_separator(QFrame.Shape.HLine))
		thanks_label = QLabel('<i>Программа создана за спасибо.<br>'
							'Более весомая благодарность стимулирет разработчика поддержать продукт :)</i>')
		thanks_label.setWordWrap(True)
		thanks_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
		info_layout.addWidget(thanks_label)

		qr_path = Path(Icons.resource_path('UI/icons/donate_qr.jpg'))
		pixmap = QPixmap(str(qr_path))
		pixmap = pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
		qr_label = QLabel()
		qr_label.setPixmap(pixmap)
		qr_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
		info_layout.addWidget(qr_label)

		info_group.setLayout(info_layout)

		main_layout.addWidget(about_group)
		main_layout.addWidget(info_group)
		main_layout.addStretch()

		# Центрируем диалог относительно главного окна
		dialog.resize(700, 450)
		dialog.setMaximumHeight(700)
		dialog.setMaximumWidth(450)
		dialog.exec()
# --------------------------------------------------------------------------------------
# =================================== ВКЛАДКА: ПРОЕКТ ==================================
# --------------------------------------------------------------------------------------

class MainProjectTab(QWidget):
	"""Первая вкладка с управлением проектом и разделами."""
	def __init__(self, main_window):
		super().__init__()
		self.main_window: Project_Window = main_window
		self.project: Project = main_window.project
		self.setup_ui()

	def setup_ui(self):
		# ----------------------------- Справка -----------------------------------------
		#help_layout = QHBoxLayout()

		btn_help = QPushButton()
		btn_help.setToolTip('Справка')
		btn_help.setIcon(Icons.help)
		btn_help.setIconSize(QSize(26, 26))
		btn_help.setFixedWidth(28)
		btn_help.setFixedHeight(28)
		btn_help.clicked.connect(self.main_window.get_help)

		# -------------------------- Группа каталога ------------------------------------
		dir_group = QGroupBox(' Каталог ')
		dir_h_box = QHBoxLayout()

		directory_text = 'Текущая директория: ' + self.main_window.get_project_dir()
		self.label_project_dir = QLabel(directory_text)
		dir_h_box.addWidget(self.label_project_dir)

		self.button_open_dir = QPushButton()
		self.button_open_dir.setToolTip('Открыть')
		self.button_open_dir.setIcon(Icons.folder_open)
		self.button_open_dir.setIconSize(QSize(22, 22))
		self.button_open_dir.setFixedWidth(28)
		self.button_open_dir.setFixedHeight(28)
		self.button_open_dir.setEnabled(False)
		self.button_open_dir.clicked.connect(self.main_window.open_project_folder)


		self.button_new_refinement = QPushButton('Копировать ревизию')
		self.button_new_refinement.setFixedWidth(156)
		self.button_new_refinement.setEnabled(False)
		self.button_new_refinement.clicked.connect(self.main_window.new_refinement)

		dir_h_box.addWidget(btn_help)
		dir_h_box.addWidget(create_separator(QFrame.Shape.VLine, QFrame.Shadow.Sunken))
		dir_h_box.addWidget(self.button_open_dir)
		dir_h_box.addWidget(self.button_new_refinement)


		button_directory = QPushButton('Задать каталог проекта')
		button_directory.setIcon(Icons.database)
		button_directory.clicked.connect(lambda: self.main_window.set_project_dir())
		button_directory.setFixedWidth(195)

		dir_h_box.addWidget(button_directory)
		dir_h_box.setAlignment(button_directory, Qt.AlignmentFlag.AlignTop)

		dir_group.setLayout(dir_h_box)
		dir_group.setMaximumHeight(70)

		# ----------------------- Группа информации о проекте  --------------------------

		self.construction_site_label = QLabel('<b>Наименование объекта:</b>' + self.get_constaction_site())
		self.construction_site_label.setWordWrap(True)
		self.construction_site_label.setMinimumHeight(50)
		self.code_label = QLabel('<b>Шифр:</b> ' + self.get_code())
		self.verifier_label = QLabel('<b>ГИП:</b> ' + self.get_verifier())
		self.description = QTextBrowser()
		self.description.setText(self.get_description())
		self.description.setMaximumHeight(100)
		self.description.setMaximumWidth(600)
		self.description.setAlignment(Qt.AlignmentFlag.AlignTop)

		project_info_subbox_left = QVBoxLayout()
		project_info_subbox_left.addWidget(self.construction_site_label)
		project_info_subbox_left.addWidget(create_separator(QFrame.Shape.HLine))
		project_info_subbox_left.addWidget(self.code_label)
		project_info_subbox_left.addWidget(self.verifier_label)
		
		project_info_subbox_median = QVBoxLayout()
		project_info_subbox_median.addWidget(QLabel('<b>Описание объекта:</b>'))
		project_info_subbox_median.addWidget(self.description)

		self.button_project_info_edit = QPushButton('Редактировать данные')
		self.button_project_info_edit.setIcon(Icons.edit)
		self.button_project_info_edit.clicked.connect(self.open_PI_edit_dialog)
		self.button_project_info_edit.setEnabled(False)
		self.button_project_info_edit.setFixedWidth(180)

		self.button_project_info_save = QPushButton('Сохранить изменения')
		self.button_project_info_save.setIcon(Icons.save)
		self.button_project_info_save.setEnabled(False)
		self.button_project_info_save.clicked.connect(self.main_window.save_project_data)
		self.button_project_info_save.setFixedWidth(180)

		project_info_subbox_right = QVBoxLayout()
		project_info_subbox_right.addWidget(self.button_project_info_edit)
		project_info_subbox_right.addWidget(self.button_project_info_save)

		project_info_box = QHBoxLayout()

		project_info_box.addLayout(project_info_subbox_left)
		project_info_box.addWidget(create_separator(QFrame.Shape.VLine))
		project_info_box.addLayout(project_info_subbox_median)
		project_info_box.addWidget(create_separator(QFrame.Shape.VLine))
		project_info_box.addLayout(project_info_subbox_right)

		project_info_group = QGroupBox(' Информация о проекте ')
		project_info_group.setLayout(project_info_box)
		project_info_group.setMaximumHeight(130)

		# ------------------------- Группа информации о разделах  -----------------------

		self.project_BoQs_table = BoQs_TableWidget(main_tab=self)
		self.project_BoQs_table.itemSelectionChanged.connect(self.update_logs_table)

		self.button_BoQ_create = QPushButton('Создать ведомость')
		self.button_BoQ_create.setIcon(Icons.note_add)
		self.button_BoQ_create.clicked.connect(self.open_BoQ_macker_dialog)
		self.button_BoQ_create.setEnabled(False)

		self.button_BoQ_open = QPushButton('Открыть ведомость')
		self.button_BoQ_open.setIcon(Icons.file_open)
		self.button_BoQ_open.clicked.connect(self.open_file_BoQ)
		self.button_BoQ_open.setEnabled(False)

		self.button_BoQ_edit = QPushButton('Редактировать ведомость')
		self.button_BoQ_edit.setIcon(Icons.edit_document)
		self.button_BoQ_edit.clicked.connect(self.edit_selected_BoQ)
		self.button_BoQ_edit.setEnabled(False)

		self.button_BoQs_reload = QPushButton('Обновить список')
		self.button_BoQs_reload.setIcon(Icons.reload)
		self.button_BoQs_reload.clicked.connect(self.reload_BoQs_table)
		self.button_BoQs_reload.setEnabled(False)

		self.button_BoQs_unlock = QPushButton('Разблокировать')
		self.button_BoQs_unlock.setIcon(Icons.unlock)
		self.button_BoQs_unlock.clicked.connect(self.unlock_BoQ)
		self.button_BoQs_unlock.setEnabled(False)

		self.button_BoQ_remove = QPushButton('Удалить ведомость')
		self.button_BoQ_remove.setIcon(Icons.delete)
		self.button_BoQ_remove.clicked.connect(self.delete_selected_BoQ)
		self.button_BoQ_remove.setEnabled(False)

		self.button_BoQs_export = QPushButton('Экспорт')
		self.button_BoQs_export.setIcon(Icons.upload)
		self.button_BoQs_export.clicked.connect(self.export_selected_BoQs)
		self.button_BoQs_export.setEnabled(False)

		self.button_BoQs_import = QPushButton('Импорт')
		self.button_BoQs_import.setIcon(Icons.download)
		self.button_BoQs_import.clicked.connect(self.import_files)
		self.button_BoQs_import.setEnabled(False)

		# Сборка панели с кнопками
		work_with_BoQ_subbox = QVBoxLayout()
		work_with_BoQ_buttons = QHBoxLayout()
		first_stack = QVBoxLayout()
		second_stack = QVBoxLayout()

		first_stack.addWidget(self.button_BoQ_create)
		first_stack.addWidget(self.button_BoQ_open)
		first_stack.addWidget(self.button_BoQ_edit)
		first_stack.addWidget(self.button_BoQs_reload)

		second_stack.addWidget(self.button_BoQs_unlock)
		second_stack.addWidget(self.button_BoQ_remove)
		second_stack.addWidget(self.button_BoQs_export)
		second_stack.addWidget(self.button_BoQs_import)

		work_with_BoQ_buttons.addLayout(first_stack)
		work_with_BoQ_buttons.addLayout(second_stack)
		work_with_BoQ_subbox.addLayout(work_with_BoQ_buttons)

		# Таблица логов
		self.table_BoQ_logs = BoQLogs_TableWidget()
		self.table_BoQ_logs.setMaximumWidth(400)
		label_BoQ_logs = QLabel('Журнал изменений ведомости:')
		work_with_BoQ_subbox.addWidget(label_BoQ_logs)
		work_with_BoQ_subbox.addWidget(self.table_BoQ_logs)

		project_BoQs_box = QHBoxLayout()
		project_BoQs_box.addWidget(self.project_BoQs_table)
		project_BoQs_box.addLayout(work_with_BoQ_subbox)

		project_BoQs_group = QGroupBox(' Ведомости ')
		project_BoQs_group.setLayout(project_BoQs_box)

		# ------------------------------- Сборка вкладки  -------------------------------
		main_layout = QVBoxLayout(self)
		main_layout.addWidget(dir_group)
		main_layout.addWidget(project_info_group)
		main_layout.addWidget(project_BoQs_group)

	# ---------------- Вспомогательные методы для получения данных из проекта -----------
	def get_constaction_site(self):
		return "Нет данных" if not self.project else self.project.construction_site

	def get_code(self):
		return "Нет данных" if not self.project else self.project.code

	def get_verifier(self):
		return "Нет данных" if not self.project else self.project.verifier['Name']
	
	def get_description(self):
		return "Нет данных" if not self.project else self.project.description

	def load_project_data(self):
		"""Обновляет лейблы информации о проекте."""
		self.construction_site_label.setText(f'<b>Наименование объекта:<br></b> {self.get_constaction_site()}')
		self.code_label.setText(f'<b>Шифр:</b> {self.get_code()}')
		self.verifier_label.setText(f'<b>ГИП:</b> {self.get_verifier()}')
		self.description.setText(f'{self.get_description().replace('\n','<br>')}')

	def tab_selected(self):
		"""Вызывается при переключении на вкладку 'Проект'."""
		if self.project is not None:
			self.reload_BoQs_table()

	# ---------------------- Методы, связанные с выбором директории ---------------------
	def on_project_dir_changed(self):
		"""Вызывается главным окном после установки директории."""
		self.project = self.main_window.project
		self.label_project_dir.setText('Текущая директория: ' + self.main_window.get_project_dir())
		self.label_project_dir.adjustSize()
		# Активируем кнопки
		self.button_open_dir.setEnabled(True)
		self.button_new_refinement.setEnabled(True)
		self.button_project_info_save.setEnabled(True)
		self.button_project_info_edit.setEnabled(True)
		self.button_BoQ_create.setEnabled(True)
		self.button_BoQ_open.setEnabled(True)
		self.button_BoQ_edit.setEnabled(True)
		self.button_BoQs_unlock.setEnabled(True)
		self.button_BoQs_reload.setEnabled(True)
		self.button_BoQ_remove.setEnabled(True)
		self.button_BoQs_export.setEnabled(True)
		self.button_BoQs_import.setEnabled(True)
		# Загружаем данные и обновляем таблицы
		self.load_project_data()
		self.reload_BoQs_table()

		# Обновляем вкладки
		self.main_window.settings_tab.set_project(self.project)


	# --------------------------- Работа с разделами ВОР --------------------------------
	def reload_BoQs_table(self):
		if self.project is None:
			return
		self.project_BoQs_table.update_table(self.project.BoQs)
		self.update_logs_table()

	def update_logs_table(self):
		if self.project is None:
			return
		current_row = self.project_BoQs_table.currentRow()
		if current_row == -1:
			self.table_BoQ_logs.clearContents()
			self.table_BoQ_logs.setRowCount(0)
			return
		data_BoQ: File_BoQ = self.project_BoQs_table.files[current_row]
		if data_BoQ:
			log_list = data_BoQ.log_list
			self.table_BoQ_logs.update_table(log_list)
		else:
			self.table_BoQ_logs.clearContents()
			self.table_BoQ_logs.setRowCount(0)

	def preliminary_BoQ_check(self, action):
		if self.project is None:
			return (False, None)
		current_row = self.project_BoQs_table.currentRow()
		if current_row == -1:
			QMessageBox.warning(self, action, "Сначала выберите раздел в таблице.")
			return (False, None)
		file_BoQ: File_BoQ = self.project_BoQs_table.files[current_row]
		try:
			status, user = file_BoQ.status_lock
			if status:
				username = user if user else 'не установлено'
				QMessageBox.warning(self, action, f'Файл заблокирован другим пользователем: {username}')
				return (False, file_BoQ)
		except KeyError:
			QMessageBox.warning(self, action, "Файл  ведомости не найден.")
			self.reload_BoQs_table()
			return (False, None)
		return (True, file_BoQ)

	def open_BoQ_macker_dialog(self):
		if self.project is None:
			return
		dialog = BoQ_Dialog(
			self,
			performers=self.project.performers,
			performers_positions=self.project.posts,
			list_BoQs=self.project.BoQs,
			project= self.project
		)
		if dialog.exec() == QDialog.DialogCode.Accepted:
			metadata = dialog.get_data()
			filename = metadata.pop('FileName')
			act = self.project.create_BoQ(filename, metadata)
			if not act:
				QMessageBox.warning(self, 'Создание раздела', 'Не удалось создать файл')
			self.reload_BoQs_table()

	def edit_selected_BoQ(self):
		if self.project is None:
			return
		access, file_BoQ = self.preliminary_BoQ_check('Редактирование раздела')
		if access:
			file_BoQ: File_BoQ
			try:
				check = file_BoQ.lock()
			except FileIsBusy as e:
				QMessageBox.warning(
					self,
					'Редактирование', 
					f"Файл уже заблокирован пользователем {e.get_data()}."
				)
				return
			except FileNotFoundError as e:
				QMessageBox.warning(
					self,
					'Редактирование', 
					f'Файл "{file_BoQ.path.name}" больше не существует.'
				)
				return
			dialog = BoQ_Dialog(
				self,
				performers=self.project.performers,
				performers_positions=self.project.posts,
				list_BoQs=self.project.BoQs,
				edit_mode=True,
				current_data=file_BoQ,
				project= self.project
			)
			result = dialog.exec()
			if result == QDialog.DialogCode.Accepted:
				file_BoQ.edit_file(dialog.get_data()) # .unlock() зашит в метод
				self.reload_BoQs_table()
			else:
				file_BoQ.unlock()

	def delete_selected_BoQ(self):
		access, file_BoQ = self.preliminary_BoQ_check('Удаление раздела')
		if access:
			reply = Requestion.ask(
				self,
				'Подтверждение удаления',
				f'Вы действительно хотите удалить раздел «{file_BoQ.path.name}»?'
			)
			if reply == QMessageBox.StandardButton.Yes:
				file_BoQ.remove_file()
				self.reload_BoQs_table()

	def open_PI_edit_dialog(self):
		if self.project is None:
			return
		dialog = Project_Info_Edit_Dialog(
			self,
			chiefs= self.project.chiefs,
			construction_site = self.project.construction_site,
			code = self.project.code,
			verifier_name = self.project.verifier['Name'],
			description = self.project.description
		)
		if dialog.exec() == QDialog.DialogCode.Accepted:
			data = dialog.get_data()
			self.project.construction_site = data['ConstructionSite']
			self.project.code = data['Code']
			self.project.verifier['Name'] = data['VerifierName']
			self.project.description = data['Description']
			self.load_project_data()
	
	def open_file_BoQ(self) -> bool:
		"""Открывает все выбранные ведомости во вкладке 'Ведомости'."""
		if self.project is None:
			return False
		
		# Собираем индексы выделенных строк
		selected_rows = set()
		for idx in self.project_BoQs_table.selectedIndexes():
			selected_rows.add(idx.row())
		if not selected_rows:
			QMessageBox.warning(self, 'Открытие файлов', "Сначала выберите ведомости в таблице.")
			return False
		
		all_files = self.project.project_BoQs
		busy_files = []

		# Сортируем для предсказуемого порядка
		for row in sorted(selected_rows):
			try:
				file_manager = self.project.open_BoQ(row)
				if file_manager.read_mode:
					name = self.project.BoQs[row].object_name
					busy_files.append(name)
				if hasattr(self.main_window, 'BoQs_tab'):
					self.main_window.BoQs_tab.add_BoQ_subtab(file_manager)
				else:
					QMessageBox.warning(self, 'Открытие файла', "Вкладка 'Разделы' не инициализирована.")
					return False
			except FileNotFoundError:
				QMessageBox.warning(self, 'Открытие файла', f"Файл ведомости не найден ({all_files[row].path.name}).")
			except Exception as e:
				QMessageBox.warning(self, 'Открытие файла', f"Не удалось открыть ведомость ({all_files[row].path.name}): {e}")

		if busy_files:
			busy_files = [f'<li>{file}</li>' for file in busy_files]
			output = f'<ul>{''.join(busy_files)}</ul>'
			QMessageBox.warning(
				self, 'Открытие файла', 
				f'<p>Файл(ы) открыт(ы) в режиме чтения:</p>{output}'
			)

		self.reload_BoQs_table()   # обновляем статусы блокировок

		
	def unlock_BoQ(self):
		if self.project is None:
			return
		current_row = self.project_BoQs_table.currentRow()
		if current_row == -1:
			QMessageBox.warning(self, 'Разблокировка файла', "Сначала выберите раздел в таблице.")
			return False
		file_BoQ: File_BoQ = self.project_BoQs_table.files[current_row]
		info1 = 'Это может привести к взаимной перезаписи файлов разными пользователями.'
		info2 = 'Используйте эту функцию только если уверены, что никто другой с этим файлом не работает'
		reply = Requestion.ask(
			self,
			'Разблокировка',
			f'Вы уверены, что  хотите снять блокировку для «{file_BoQ.path.name}»?\n{info1}\n{info2}'
		)
		
		if reply == QMessageBox.StandardButton.Yes:
			file_BoQ.remove_lock()
			self.reload_BoQs_table()
	
	def import_files(self):
		if not self.project:
			return

		# Диалог выбора файлов с расширенным фильтром
		files, _ = QFileDialog.getOpenFileNames(
			self,
			"Выберите файлы для импорта (Excel или XML)",
			"",
			"Поддерживаемые форматы (*.xlsx *.xlsm *.xml *.gge);;Excel файлы (*.xlsx *.xlsm);;XML файлы (*.xml *.gge)"
		)
		if not files:
			return

		convertor = Convertor(self.project)
		success_count = 0
		total_items = 0

		for file_path_str in files:
			file_path = Path(file_path_str)
			suffix = file_path.suffix.lower()

			if suffix in ('.xml', '.gge'):
				# Импорт из XML
				try:
					convertor.import_from_xml_3p01(file_path)
					success_count += 1
					self.log_message(f"Импорт из XML: {file_path.name} успешно завершён")
				except ShemaError as e:
					QMessageBox.warning(self, 'Импорт XML', e.info())
				except Exception as e:
					QMessageBox.warning(self, 'Импорт XML', f"Ошибка при импорте {file_path.name}:\n{e}")
			elif suffix in ('.xlsx', '.xlsm'):
				# Импорт из Excel
				try:
					# Проверяем, что файл можно открыть и получить листы
					wb = load_workbook(file_path, data_only=True, read_only=True)
					sheet_names = wb.sheetnames
					wb.close()
				except Exception as e:
					QMessageBox.warning(self, 'Импорт Excel', f"Не удалось открыть файл {file_path.name}:\n{e}")
					continue

				if not sheet_names:
					QMessageBox.warning(self, 'Импорт Excel', f"Файл {file_path.name} не содержит листов.")
					continue

				# Показываем диалог выбора листов
				selected_sheets = self.show_sheet_selection_dialog(file_path.name, sheet_names)
				if not selected_sheets:
					continue  # пользователь отменил выбор

				# Запускаем импорт выбранных листов
				try:
					processed = convertor.import_from_excel(file_path, sheet_names=selected_sheets)
					if processed > 0:
						success_count += 1
						total_items += processed
					else:
						QMessageBox.warning(self, 'Импорт Excel', f"В файле {file_path.name} не найдено подходящих листов (A1='Документ').")
				except Exception as e:
					QMessageBox.warning(self, 'Импорт Excel', f"Ошибка при импорте {file_path.name}:\n{e}")
			else:
				QMessageBox.warning(self, 'Импорт', f"Неподдерживаемый формат файла: {suffix}")

		# Обновляем таблицу разделов после всех импортов
		self.reload_BoQs_table()

		# Итоговое сообщение
		if success_count > 0:
			if total_items:
				QMessageBox.information(self, 'Импорт завершён',
					f"Импортировано файлов: {success_count}\nВсего листов/разделов: {total_items}")
			else:
				QMessageBox.information(self, 'Импорт завершён',
					f"Успешно импортировано файлов: {success_count}")
		elif success_count == 0 and files:
			QMessageBox.warning(self, 'Импорт', "Ни один файл не был импортирован.")

	def show_sheet_selection_dialog(self, file_name, sheet_names):
		"""
		Отображает диалог с множественным выбором листов.
		Возвращает список выбранных имён листов или None, если диалог отменён.
		"""
		dialog = QDialog(self)
		dialog.setWindowTitle(f"Выбор листов для импорта: {file_name}")
		dialog.setModal(True)
		layout = QVBoxLayout(dialog)

		label = QLabel("Выберите листы для конвертации (можно несколько):")
		layout.addWidget(label)

		list_widget = QListWidget()
		list_widget.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
		for name in sheet_names:
			list_widget.addItem(name)
		layout.addWidget(list_widget)

		buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
		buttons.accepted.connect(dialog.accept)
		buttons.rejected.connect(dialog.reject)
		layout.addWidget(buttons)

		if dialog.exec() != QDialog.DialogCode.Accepted:
			return None

		selected = [item.text() for item in list_widget.selectedItems()]
		return selected

	def log_message(self, msg):
		"""Выводит сообщение в лог (если есть виджет лога, иначе в консоль)."""
		# В текущем интерфейсе нет виджета лога на главной вкладке, поэтому просто print
		print(msg)

	def export_selected_BoQs(self):
		"""Экспорт выбранных разделов в XML/GGE/PDF"""
		if self.project is None:
			QMessageBox.warning(self, "Экспорт", "Проект не загружен.")
			return

		selected_rows = set(self.project_BoQs_table.selectedIndexes())
		if not selected_rows:
			QMessageBox.warning(self, "Экспорт", "Сначала выберите ведомости в таблице.")
			return

		rows = {index.row() for index in selected_rows}
		selected_files = [self.project_BoQs_table.files[row] for row in rows]

		# 1. Проверка блокировок
		available_files = []
		locked_files = []
		for file_BoQ in selected_files:
			try:
				status, user = file_BoQ.status_lock
				if status:
					locked_files.append((file_BoQ.path.name, user))
				else:
					available_files.append(file_BoQ)
			except KeyError:
				locked_files.append((file_BoQ.path.name, "не определён"))

		if locked_files:
			msg = "Следующие ведомости заблокированы и не могут быть экспортированы:\n"
			for name, user in locked_files:
				msg += f"- {name} (пользователь: {user})\n"
			QMessageBox.warning(self, "Экспорт", msg)
			if not available_files:
				return

		# 2. Предварительная валидация и сбор файлов, готовых к экспорту
		export_list = []  # (file_BoQ, manager) или просто file_BoQ
		problems_map = {}  # file_name -> список проблем

		for file_BoQ in available_files:
			# Загружаем менеджер (без блокировки, только чтение)
			try:
				manager = self.project.open_BoQ_for_export(file_BoQ)
			except Exception as e:
				QMessageBox.warning(self, "Экспорт", f"{file_BoQ.path.name}: не удалось загрузить - {e}")
				continue

			# Проверка валидности
			data = manager.data_validation
			invalid_quantities = data.get('invalid_quantities')
			incorrect_positions = data.get('incorrect_positions')
			error_link_positions = data.get('error_link_positions')
			nonlink_positions = data.get('nonlink_positions')

			if invalid_quantities:
				QMessageBox.warning(self, "Экспорт", f"{file_BoQ.path.name}: невалидные значения объёмов. Экспорт отменён.")
				continue

			problems = []
			if incorrect_positions:
				problems.append('не все позиции имеют статус готовности')
			if nonlink_positions:
				problems.append('есть позиции с пустыми ссылками (будет создана заглушка)')
			if error_link_positions:
				problems.append('некоторые ссылки некорректны (результат может привести к ошибке)')

			if problems:
				problems_map[file_BoQ.path.name] = problems
				reply = Requestion.ask(
					self,
					'Предупреждение',
					f"В разделе «{file_BoQ.path.name}» есть проблемы:\n" + "\n".join(problems) + "\nПродолжить экспорт?",
					with_cancel = False
				)
				if reply == QMessageBox.StandardButton.No:
					continue

			export_list.append((file_BoQ, manager))

		if not export_list:
			QMessageBox.information(self, "Экспорт", "Нет разделов для экспорта.")
			return

		# 3. Выбор формата экспорта
		exp_dialog = Export_Dialog(self)
		if exp_dialog.exec() != QDialog.DialogCode.Accepted:
			return
		form = exp_dialog.get_data()

		# 4. Прогресс-диалог (теперь без прерываний)
		progress = QProgressDialog("Экспорт разделов...", "Отмена", 0, len(export_list), self)
		progress.setWindowModality(Qt.WindowModality.WindowModal)
		progress.setMinimumDuration(0)
		progress.setValue(0)

		success_count = 0
		errors = []
		convertor = Convertor(self.project)

		for idx, (file_BoQ, manager) in enumerate(export_list):
			if progress.wasCanceled():
				break
			progress.setLabelText(f"Экспорт: {file_BoQ.path.name}")
			progress.setValue(idx)
			QApplication.processEvents()  # обновляем UI

			try:
				if form[0] in ('.gge', '.xml'):
					convertor.create_xml_3p01(manager, form[0])
					success_count += 1
				else:
					convertor.export_to_pdf(manager, form[1])
					success_count += 1
			except Exception as e:
				errors.append(f"{file_BoQ.path.name}: ошибка экспорта - {e}")

		progress.setValue(len(export_list))
		progress.close()

		# Итоговое сообщение
		msg = f"Экспорт завершён.\nУспешно: {success_count} из {len(export_list)}"
		if errors:
			msg += "\n\nОшибки:\n" + "\n".join(errors[:10])
			if len(errors) > 10:
				msg += f"\n... и ещё {len(errors)-10} ошибок"
			QMessageBox.warning(self, "Экспорт", msg)
		else:
			QMessageBox.information(self, "Экспорт", msg)

		if success_count:
			match form[0]:
				case '.gge' | '.xml':
					self.project.open_xml_folder()
				case '.pdf':
					self.project.open_pdf_folder()
		

class Project_Info_Edit_Dialog(QDialog):
	""" Окно редактирования данных проекта """
	def __init__(self, parent, chiefs, construction_site, code, verifier_name, description):
		super().__init__(parent)
		self.setWindowTitle('Редактирование информации о проекте')
		self.setModal(True)

		edit_layout = QVBoxLayout(self)

		self._construction_site_edit_line = QPlainTextEdit(construction_site)
		self._construction_site_edit_line.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth)
		self._construction_site_edit_line.setMaximumHeight(100)
		self._construction_site_edit_line.setMinimumWidth(600)
		edit_layout.addWidget(QLabel('Наименование объекта:'))
		edit_layout.addWidget(self._construction_site_edit_line)

		self._code_edit_line = QLineEdit(code)
		edit_layout.addWidget(QLabel('Шифр: '))
		edit_layout.addWidget(self._code_edit_line)

		self._verifier_edit_combobox = QComboBox()
		verifiers_list = chiefs
		self._verifier_edit_combobox.addItems(verifiers_list)
		self._verifier_edit_combobox.setEditable(True)
		self._verifier_edit_combobox.setCurrentText(verifier_name)
		edit_layout.addWidget(QLabel('ГИП: '))
		edit_layout.addWidget(self._verifier_edit_combobox)
		

		self.description_edit_plain = QPlainTextEdit(description)
		self.description_edit_plain.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth)
		self.description_edit_plain.setMinimumWidth(600)
		edit_layout.addWidget(QLabel('Описание объекта:'))
		edit_layout.addWidget(self.description_edit_plain)

		project_info_edit_buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
		ok_button = project_info_edit_buttons.button(QDialogButtonBox.StandardButton.Ok)
		ok_button.setText('Применить')
		cancel_button = project_info_edit_buttons.button(QDialogButtonBox.StandardButton.Cancel)
		cancel_button.setText('Отменить')
		project_info_edit_buttons.accepted.connect(self.accept)
		project_info_edit_buttons.rejected.connect(self.reject)

		edit_layout.addWidget(project_info_edit_buttons)

	def get_data(self):
		output = {
			'ConstructionSite': self._construction_site_edit_line.toPlainText(),
			'Code': self._code_edit_line.text(),
			'VerifierName': self._verifier_edit_combobox.currentText(),
			'Description': self.description_edit_plain.toPlainText()
		}
		return output



class BoQs_TableWidget(QTableWidget):
	"""
	Таблица для отображения списка разделов.
	"""
	def __init__(self, parent=None, main_tab=None):
		super().__init__(parent)
		self.main_tab = main_tab
		self.setColumnCount(9)
		self.setHorizontalHeaderLabels([
			"Номер ВОР", "Наименование", "Статус",
			"Исполнитель", "Должность", "Последняя Дата",
			"Доступность", "№ ЛСР", "Файл"])
		self.files = []
		for i in range(self.columnCount()):
			self.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
		self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
		self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
		self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
		self.customContextMenuRequested.connect(self.on_table_context_menu)


	def update_table(self, list_BoQs):
		self.files = list_BoQs
		self.setRowCount(len(list_BoQs))
		for row, obj in enumerate(list_BoQs):
			obj: File_BoQ
			self.setItem(row, 0, QTableWidgetItem(obj.num))
			self.setItem(row, 1, QTableWidgetItem(obj.object_name))
			status = "Готово" if obj.status_done else "Не готово"
			status_item = QTableWidgetItem(status)
			status_color = "#38A14F" if status == 'Готово' else "#C22222"
			status_item.setForeground(QColor(status_color)) 
			status_item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter)
			self.setItem(row, 2, status_item)
			self.setItem(row, 3, QTableWidgetItem(obj.composer))
			self.setItem(row, 4, QTableWidgetItem(obj.composer_position))
			self.setItem(row, 5, QTableWidgetItem(obj.date))
			if not obj.is_locked:
				access = 'Доступно'
				access_color = "#38A14F"
			else:
				access = f'Открыт пользователем {obj.active_user}'
				access_color = "#C22222"
			access_item = QTableWidgetItem(access)
			access_item.setForeground(QColor(access_color))
			access_item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter)
			self.setItem(row, 6, access_item)
			self.setItem(row, 7, QTableWidgetItem(obj.local_estimate))
			self.setItem(row, 8, QTableWidgetItem(obj.path.name))

	def on_table_context_menu(self, pos):
		# Получаем позицию и индекс ячейки
		index = self.indexAt(pos)
		if not index.isValid():
			return
		row = index.row()
		# Проверяем, что есть выделенные строки
		if self.rowCount() == 0:
			return
		# Создаём меню
		menu = QMenu(self)

		# Действия, которые всегда доступны (если есть выделение)
		action_open = menu.addAction("Открыть")
		action_open.triggered.connect(self.main_tab.open_file_BoQ)

		action_edit = menu.addAction("Редактировать")
		action_edit.triggered.connect(self.main_tab.edit_selected_BoQ)

		action_delete = menu.addAction("Удалить")
		action_delete.triggered.connect(self.main_tab.delete_selected_BoQ)

		menu.addSeparator()

		action_unlock = menu.addAction("Разблокировать")
		action_unlock.triggered.connect(self.main_tab.unlock_BoQ)

		action_export = menu.addAction("Экспорт")
		action_export.triggered.connect(self.main_tab.export_selected_BoQs)

		menu.addSeparator()

		action_reload = menu.addAction("Обновить список")
		action_reload.triggered.connect(self.main_tab.reload_BoQs_table)

		# Показываем меню в позиции курсора
		menu.exec(self.viewport().mapToGlobal(pos))		

class BoQLogs_TableWidget(QTableWidget):
	"""
	Таблица для отображения реестра изменений текущего выделенного раздела.
	"""
	def __init__(self, parent=None):
		super().__init__(parent)
		self.setColumnCount(2)
		self.setHorizontalHeaderLabels(["Дата", "Событие"])
		self.setColumnWidth(0, 100)
		self.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
		self.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
		self.setWordWrap(True)
		self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
		self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
		self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

	def update_table(self, log_list):
		self.setRowCount(len(log_list))
		for row, entry in enumerate(log_list):
			self.setItem(row, 0, QTableWidgetItem(entry.get('Date')))
			self.setItem(row, 1, QTableWidgetItem(entry.get('Event')))
		self.resizeRowsToContents()



class BoQ_Dialog(QDialog):
	"""Окно создания/редактирования раздела ВОР"""
	def __init__(self, parent, performers, performers_positions, list_BoQs,
				 edit_mode=False, current_data: File_BoQ | None =None, project = None):
		super().__init__(parent)
		window_name = 'Редактирование раздела ВОР' if edit_mode else 'Создание нового раздела ВОР'
		self.setWindowTitle(window_name)
		self.setModal(True)

		self._template = None
		self.list_BoQs = list_BoQs
		self.edit_mode = edit_mode
		self.project = project
		self.current_data = current_data

		edit_layout = QHBoxLayout(self)
		left_layout = QVBoxLayout()
		right_layout = QVBoxLayout()

		# Поля ввода
		__file_name_edit = current_data.path.stem if self.edit_mode else '00_Наименование работы'
		__template_BoQs = (
			'01_Подготовительные_работы',
			'02_Земляное_полотно',
			'03_Дорожная_одежда',
			'04_Примыкания_и_пересечения',
			'05_Автобусные_остановки',
			'06_Сопряжения',
			'07_Водоотвод',
			'08_Обустройство',
			'09_Площадка_для_ВЗиС',
			'10_Крайние_опоры',
			'11_Промежуточные_опоры',
			'12_Пролетные_строения',
			'13_Мостовое_полотно',
			'14_Сопряжение_моста_с_насыпью',
			'15_Регуляционные_сооружения',
			'16_Устройство_объездного_моста',
			'17_Устройство_объездной_дороги',
			'18_Разборка_объездного_моста',
			'19_Разборка_объездной_дороги',
			'20_СВСиУ_Рабочие_площадки_Рабочий_мост',
		)
		self._file_name_combobox = QComboBox()
		self._file_name_combobox.addItems(__template_BoQs)
		self._file_name_combobox.setEditable(True)
		self._file_name_combobox.setCurrentText(__file_name_edit)
		self._file_name_combobox.currentTextChanged.connect(self.update_object_name)
		left_layout.addWidget(QLabel('Наименование файла раздела:'))
		left_layout.addWidget(self._file_name_combobox)

		__current_object_name = current_data.object_name if self.edit_mode else ''
		self._object_name_plain = QPlainTextEdit(__current_object_name)
		self._object_name_plain.setPlaceholderText('Наименование раздела работы')
		self._object_name_plain.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth)
		self._object_name_plain.setMaximumHeight(45)

		left_layout.addWidget(QLabel('Наименование раздела:'))
		left_layout.addWidget(self._object_name_plain)

		__current_num = current_data.num if self.edit_mode else ''
		self._num_edit_line = QLineEdit(__current_num)
		self._num_edit_line.setPlaceholderText('ВО-0')
		left_layout.addWidget(QLabel('Номер ведомости:'))
		left_layout.addWidget(self._num_edit_line)

		# Комбобоксы с автодополнением
		performers_list = performers
		performers.sort()
		self._composer_combo_box = QComboBox()
		self._composer_combo_box.addItems(performers_list)
		self._composer_combo_box.setEditable(True)
		completer = QCompleter(performers_list)
		self._composer_combo_box.setCompleter(completer)
		left_layout.addWidget(QLabel('Составил:'))
		left_layout.addWidget(self._composer_combo_box)

		positions_list = performers_positions
		self._composer_positions_combo_box = QComboBox()
		self._composer_positions_combo_box.addItems(positions_list)
		self._composer_positions_combo_box.setEditable(True)
		completer_positions = QCompleter(positions_list)
		self._composer_positions_combo_box.setCompleter(completer_positions)
		if self.edit_mode:
			self._composer_combo_box.setCurrentText(current_data.composer)
			self._composer_positions_combo_box.setCurrentText(current_data.composer_position)
		else:
			self._composer_positions_combo_box.setCurrentText('Ведущий инженер')
		left_layout.addWidget(QLabel('Должность:'))
		left_layout.addWidget(self._composer_positions_combo_box)

		# Дополнительные поля режима редактирования
		if self.edit_mode:
			__current_date = current_data.date
			self._date_edit_line = QLineEdit(__current_date)
			left_layout.addWidget(QLabel('Последняя дата документа:'))
			left_layout.addWidget(self._date_edit_line)

			__current_local_estimate = current_data.local_estimate
			self._local_estimate_edit_line = QLineEdit(__current_local_estimate)
			left_layout.addWidget(QLabel('№ ЛСР:'))
			left_layout.addWidget(self._local_estimate_edit_line)

			# Таблица журнала изменений
			right_layout.addWidget(QLabel('Журнал изменений:'))

			self.logs_table = QTableWidget()
			self.logs_table.setColumnCount(2)
			self.logs_table.setHorizontalHeaderLabels(['Дата', 'Событие'])
			self.logs_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
			self.logs_table.setWordWrap(True)
			self.logs_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed)
			self.logs_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
			self.logs_table.setItemDelegateForColumn(1, MultilineTextDelegate())

			log_list = current_data.log_list
			self.logs_table.setRowCount(len(log_list))
			for row, entry in enumerate(log_list):
				self.logs_table.setItem(row, 0, QTableWidgetItem(entry.get('Date', '')))
				self.logs_table.setItem(row, 1, QTableWidgetItem(entry.get('Event', '')))
			self.logs_table.resizeRowsToContents()

			right_layout.addWidget(self.logs_table)

			# Кнопки управления журналом
			logs_buttons_layout = QHBoxLayout()
			self.btn_add_log = QPushButton('Добавить запись')
			self.btn_add_log.clicked.connect(self.add_log_row)
			self.btn_remove_log = QPushButton('Удалить запись')
			self.btn_remove_log.clicked.connect(self.remove_log_row)
			logs_buttons_layout.addWidget(self.btn_add_log)
			logs_buttons_layout.addWidget(self.btn_remove_log)
			right_layout.addLayout(logs_buttons_layout)

		# Кнопки
		new_BoQ_buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
		ok_button = new_BoQ_buttons.button(QDialogButtonBox.StandardButton.Ok)
		ok_button_name = 'Применить' if self.edit_mode else 'Создать'
		ok_button.setText(ok_button_name)
		cancel_button = new_BoQ_buttons.button(QDialogButtonBox.StandardButton.Cancel)
		cancel_button.setText('Отменить')
		new_BoQ_buttons.accepted.connect(self.accept)
		new_BoQ_buttons.rejected.connect(self.reject)

		left_layout.addWidget(new_BoQ_buttons)

		# Контейнеры для ограничения ширины
		left_container = QWidget()
		left_container.setLayout(left_layout)
		if edit_mode:
			left_container.setMaximumWidth(300)

		right_container = QWidget()
		right_container.setLayout(right_layout)
		right_container.setMinimumWidth(500)

		edit_layout.addWidget(left_container)
		if edit_mode:
			edit_layout.addWidget(right_container)
			edit_layout.setStretchFactor(left_container, 1)
			edit_layout.setStretchFactor(right_container, 2)

	# ------------------------- Методы обновления наименовния -------------------------

	def update_object_name(self):
		""" Обновляет имя раздела при выборе имени файла, если оно пустое """
		non_list = {'',' ','-', None, 'None','...', '.'}
		if self._object_name_plain.toPlainText().strip() in non_list and not self.edit_mode:
			filename = text_after(self._file_name_combobox.currentText(),'_').replace('_',' ')
			self._object_name_plain.setPlainText(filename)
		
		if self._num_edit_line.text() in non_list and not self.edit_mode:
			num = text_before(self._file_name_combobox.currentText(),'_')
			num = num[1:] if num.startswith('0') else num
			self._num_edit_line.setText(f'ВО-{num}')
		
	

	# ------------------------ Методы для управления журналом -------------------------
	def add_log_row(self):
		row = self.logs_table.rowCount()
		self.logs_table.insertRow(row)
		self.logs_table.setItem(row, 0, QTableWidgetItem(self.project.now.strftime("%d.%m.%Y")))
		self.logs_table.setItem(row, 1, QTableWidgetItem(''))

	def remove_log_row(self):
		current_row = self.logs_table.currentRow()
		if current_row >= 0:
			self.logs_table.removeRow(current_row)

	# ---------------------------- Методы основных действий ---------------------------
	def accept(self):
		filename = self._file_name_combobox.currentText().strip()
		filenames = [f.path.stem for f in self.list_BoQs]
		if not filename:
			QMessageBox.warning(self, "Ошибка", "Имя файла не может быть пустым.")
			return
		if filename in filenames and not self.edit_mode:
			QMessageBox.information(self, "Внимание!", "Файл с таким именем уже существует.")
			return
		super().accept()


	def get_data(self):
		date = self._date_edit_line.text() if self.edit_mode else self.project.now.strftime("%d.%m.%Y")
		local_estimate = self._local_estimate_edit_line.text() if self.edit_mode else 'Укажите номер ЛСР'
		log_list = []

		if self.edit_mode:
			for row in range(self.logs_table.rowCount()):
				date_item = self.logs_table.item(row, 0)
				event_item = self.logs_table.item(row, 1)
				log_list.append({
					'Date': date_item.text() if date_item else '',
					'Event': event_item.text() if event_item else ''
				})

		output = {
			'FileName': self._file_name_combobox.currentText().strip(),
			'metadata': {
				'ObjectName': self._object_name_plain.toPlainText(),
				'Num': self._num_edit_line.text(),
				'Date': date,
				'Signatures': {
					'Composer': self._composer_combo_box.currentText(),
					'Composer_Position': self._composer_positions_combo_box.currentText()
				},
				'Status_Done': self.current_data.status_done if self.edit_mode else False,
				'local_estimate': local_estimate,
				'log_list': log_list,
				'note': self.current_data.note if self.edit_mode else ''
			}
		}
		return output if not self._template else self._template

class Export_Dialog(QDialog):
	def __init__(self, parent=None):
		super().__init__(parent)
		self.setWindowTitle("Экспорт ведомости")
		self.setModal(True)

		layout = QVBoxLayout(self)

		layout.addWidget(QLabel("Выберите формат экспорта:"))

		self.format_combo = QComboBox()
		self.format_combo.addItems(["XML", "GGE", "PDF по форме 1", "PDF по форме 2"])
		layout.addWidget(self.format_combo)

		btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
		btn_box.accepted.connect(self.accept)
		btn_box.rejected.connect(self.reject)
		layout.addWidget(btn_box)

	def get_data(self):
		"""Возвращает расширение файла: .xml или .gge"""
		form = self.format_combo.currentText()
		if form == "XML":
			return (".xml", None)
		elif form == "GGE":
			return (".gge", None)
		elif form == "PDF по форме 1":
			return (".pdf", 0)
		else:
			return (".pdf", 1)